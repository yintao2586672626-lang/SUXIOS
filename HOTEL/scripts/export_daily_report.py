#!/usr/bin/env python3
"""
日报表导出脚本
按照模板格式生成Excel报表
"""

import sys
import json
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
import io
import base64


def get_column_letter_by_index(index):
    """根据索引获取列字母（1-based）"""
    return get_column_letter(index)


def create_excel_report(reports_data, month_task_data=None):
    """
    按照模板格式创建日报表Excel
    
    Args:
        reports_data: 日报表数据列表
        month_task_data: 月任务数据
    
    Returns:
        bytes: Excel文件内容
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "日报表"
    
    # 默认行高
    ws.default_row_height = 25
    
    # 定义样式
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 颜色
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    blue_fill = PatternFill(start_color='B2CFEA', end_color='B2CFEA', fill_type='solid')
    green_fill = PatternFill(start_color='EBF1DE', end_color='EBF1DE', fill_type='solid')
    pink_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # 字体
    bold_font_12 = Font(bold=True, name='Arial', size=12)
    bold_font_11 = Font(bold=True, name='Arial', size=11)
    
    # 星期数组
    week_array = ['星期日', '星期一', '星期二', '星期三', '星期四', '星期五', '星期六']
    
    # ==================== 创建表头 ====================
    
    # 第1-3行：多层表头
    # 列索引映射（1-based）
    # A=1, B=2, ... Z=26, AA=27, AB=28, ...
    
    # ========== 第一行表头 ==========
    # A1: 日期（合并A1:A3）
    ws['A1'] = '日期'
    ws.merge_cells('A1:A3')
    
    # B1: 星期（合并B1:B3）
    ws['B1'] = '星期'
    ws.merge_cells('B1:B3')
    
    # A4: 合计（合并A4:B4）
    ws['A4'] = '合计'
    ws.merge_cells('A4:B4')
    
    # C1: 合计（合并C1:O2）
    ws['C1'] = '合计'
    ws.merge_cells('C1:O2')
    
    # P1: 线上客房收入（合并P1:AH1）
    ws['P1'] = '线上客房收入'
    ws.merge_cells('P1:AH1')
    
    # AI1: 线下客房收入（合并AI1:BE1）
    ws['AI1'] = '线下客房收入'
    ws.merge_cells('AI1:BE1')
    
    # BF1: 钟点房（合并BF1:BG2）
    ws['BF1'] = '钟点房'
    ws.merge_cells('BF1:BG2')
    
    # BH1: 其他收入合计（合并BH1:BN1）
    ws['BH1'] = '其他收入合计'
    ws.merge_cells('BH1:BN1')
    
    # BO1: 好评（合并BO1:BW1）
    ws['BO1'] = '好评'
    ws.merge_cells('BO1:BW1')
    
    # BX1: 流量（合并BX1:CD1）
    ws['BX1'] = '流量'
    ws.merge_cells('BX1:CD1')
    
    # CE1: 私域流量（合并CE1:CG2）
    ws['CE1'] = '私域流量'
    ws.merge_cells('CE1:CG2')
    
    # CH1: 私域订单（合并CH1:CI2）
    ws['CH1'] = '私域订单'
    ws.merge_cells('CH1:CI2')
    
    # CJ1: 会员（合并CJ1:CL2）
    ws['CJ1'] = '会员'
    ws.merge_cells('CJ1:CL2')
    
    # ========== 第二行表头 ==========
    # C2已合并到C1
    
    # P2: 线上合计（合并P2:T2）
    ws['P2'] = '线上合计'
    ws.merge_cells('P2:T2')
    
    # U2: 携程（合并U2:V2）
    ws['U2'] = '携程'
    ws.merge_cells('U2:V2')
    
    # W2: 美团（合并W2:X2）
    ws['W2'] = '美团'
    ws.merge_cells('W2:X2')
    
    # Y2: 飞猪（合并Y2:Z2）
    ws['Y2'] = '飞猪'
    ws.merge_cells('Y2:Z2')
    
    # AA2: booking（合并AA2:AB2）
    ws['AA2'] = 'booking'
    ws.merge_cells('AA2:AB2')
    
    # AC2: 抖音（合并AC2:AD2）
    ws['AC2'] = '抖音'
    ws.merge_cells('AC2:AD2')
    
    # AE2: 其他2（合并AE2:AF2）
    ws['AE2'] = '其他2'
    ws.merge_cells('AE2:AF2')
    
    # AG2: 其他3（合并AG2:AH2）
    ws['AG2'] = '其他3'
    ws.merge_cells('AG2:AH2')
    
    # AI2: 线下合计（合并AI2:AM2）
    ws['AI2'] = '线下合计'
    ws.merge_cells('AI2:AM2')
    
    # AN2: 散客（合并AN2:AO2）
    ws['AN2'] = '散客'
    ws.merge_cells('AN2:AO2')
    
    # AP2: 会员体验（合并AP2:AQ2）
    ws['AP2'] = '会员体验'
    ws.merge_cells('AP2:AQ2')
    
    # AR2: 网络体验（合并AR2:AS2）
    ws['AR2'] = '网络体验'
    ws.merge_cells('AR2:AS2')
    
    # AT2: 团队（合并AT2:AU2）
    ws['AT2'] = '团队'
    ws.merge_cells('AT2:AU2')
    
    # AV2: 协议客户（合并AV2:AW2）
    ws['AV2'] = '协议客户'
    ws.merge_cells('AV2:AW2')
    
    # AX2: 微信（合并AX2:AY2）
    ws['AX2'] = '微信'
    ws.merge_cells('AX2:AY2')
    
    # AZ2: 免费房（合并AZ2:BA2）
    ws['AZ2'] = '免费房'
    ws.merge_cells('AZ2:BA2')
    
    # BB2: 集团金卡（合并BB2:BC2）
    ws['BB2'] = '集团金卡'
    ws.merge_cells('BB2:BC2')
    
    # BD2: 集团黑金卡（合并BD2:BE2）
    ws['BD2'] = '集团黑金卡'
    ws.merge_cells('BD2:BE2')
    
    # BH2: 其他收入合计（合并BH2:BH3）
    ws['BH2'] = '其他收入合计'
    ws.merge_cells('BH2:BH3')
    
    # BI2: 停车费（合并BI2:BI3）
    ws['BI2'] = '停车费'
    ws.merge_cells('BI2:BI3')
    
    # BJ2: 餐饮（合并BJ2:BJ3）
    ws['BJ2'] = '餐饮'
    ws.merge_cells('BJ2:BJ3')
    
    # BK2: 会议活动（合并BK2:BK3）
    ws['BK2'] = '会议活动'
    ws.merge_cells('BK2:BK3')
    
    # BL2: 商品（合并BL2:BL3）
    ws['BL2'] = '商品'
    ws.merge_cells('BL2:BL3')
    
    # BM2: 会员卡费收入（合并BM2:BM3）
    ws['BM2'] = '会员卡费收入'
    ws.merge_cells('BM2:BM3')
    
    # BN2: 其他（合并BN2:BN3）
    ws['BN2'] = '其他'
    ws.merge_cells('BN2:BN3')
    
    # BO2: 好评合计（合并BO2:BQ2）
    ws['BO2'] = '好评合计'
    ws.merge_cells('BO2:BQ2')
    
    # BR2: 携程、艺龙、去哪儿（合并BR2:BS2）
    ws['BR2'] = '携程、艺龙、去哪儿'
    ws.merge_cells('BR2:BS2')
    
    # BT2: 美团（合并BT2:BU2）
    ws['BT2'] = '美团'
    ws.merge_cells('BT2:BU2')
    
    # BV2: 飞猪（合并BV2:BW2）
    ws['BV2'] = '飞猪'
    ws.merge_cells('BV2:BW2')
    
    # BX2: 携程（合并BX2:CA2）
    ws['BX2'] = '携程'
    ws.merge_cells('BX2:CA2')
    
    # CB2: 美团（合并CB2:CD2）
    ws['CB2'] = '美团'
    ws.merge_cells('CB2:CD2')
    
    # ========== 第三行表头 ==========
    # 合计区域
    headers_row3 = {
        'C3': '月目标', 'D3': '总目标完成率', 'E3': '总收入', 'F3': '客房收入',
        'G3': '客房Revpar', 'H3': '出租率', 'I3': '平均房价', 'J3': '过夜出租率',
        'K3': '过夜均价', 'L3': '过夜Revpar', 'M3': 'OTA间夜占比', 'N3': '售房数',
        'O3': '可售房数量',
        # 线上合计
        'P3': '线上目标', 'Q3': '线上目标完成率', 'R3': '总收入', 'S3': '售房数', 'T3': '平均房价',
        # 携程
        'U3': '收入', 'V3': '间夜',
        # 美团
        'W3': '收入', 'X3': '间夜',
        # 飞猪
        'Y3': '收入', 'Z3': '间夜',
        # booking
        'AA3': '收入', 'AB3': '间夜',
        # 抖音
        'AC3': '收入', 'AD3': '间夜',
        # 其他2
        'AE3': '收入', 'AF3': '间夜',
        # 其他3
        'AG3': '收入', 'AH3': '间夜',
        # 线下合计
        'AI3': '线下目标', 'AJ3': '线下目标完成率', 'AK3': '总收入', 'AL3': '售房数', 'AM3': '平均房价',
        # 散客
        'AN3': '收入', 'AO3': '间夜',
        # 会员体验
        'AP3': '收入', 'AQ3': '间夜',
        # 网络体验
        'AR3': '收入', 'AS3': '间夜',
        # 团队
        'AT3': '收入', 'AU3': '间夜',
        # 协议客户
        'AV3': '收入', 'AW3': '间夜',
        # 微信
        'AX3': '收入', 'AY3': '间夜',
        # 免费房
        'AZ3': '收入', 'BA3': '间夜',
        # 集团金卡
        'BB3': '收入', 'BC3': '间夜',
        # 集团黑金卡
        'BD3': '收入', 'BE3': '间夜',
        # 钟点房
        'BF3': '收入', 'BG3': '间夜',
        # 好评合计
        'BO3': '订单数', 'BP3': '5分点评数', 'BQ3': '好转化评率',
        # 携程、艺龙、去哪儿
        'BR3': '可评价订单数', 'BS3': '5分点评数量',
        # 美团
        'BT3': '可评价订单数', 'BU3': '5分点评数量',
        # 飞猪
        'BV3': '可评价订单数', 'BW3': '5分点评数量',
        # 携程流量
        'BX3': '列表页曝光量', 'BY3': '曝光转化率', 'BZ3': '下单转化率', 'CA3': '成交转化率',
        # 美团流量
        'CB3': '曝光人数', 'CC3': '点击率', 'CD3': '支付转化率',
        # 私域流量
        'CE3': '微信加粉率', 'CF3': '微信加粉人数', 'CG3': '新增会员人数',
        # 私域订单
        'CH3': '收入', 'CI3': '间夜',
        # 会员
        'CJ3': '售卡量', 'CK3': '售卡收入', 'CL3': '储值',
    }
    
    for cell, value in headers_row3.items():
        ws[cell] = value
    
    # 应用表头样式
    for row in range(1, 4):
        for col in range(1, 91):  # A到CL列
            cell = ws.cell(row=row, column=col)
            cell.alignment = center_align
            cell.border = thin_border
    
    ws['A1'].alignment = center_align
    ws['B1'].alignment = center_align
    
    # 设置字体
    for row in range(1, 3):
        for col in range(1, 91):
            ws.cell(row=row, column=col).font = bold_font_12
    
    for col in range(1, 91):
        ws.cell(row=3, column=col).font = bold_font_11
    
    # ==================== 第4行合计行 ====================
    data_count = len(reports_data)
    last_row = 4 + data_count  # 数据最后一行
    
    # 合计行黄色背景
    for col in range(1, 91):
        cell = ws.cell(row=4, column=col)
        cell.fill = yellow_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # 合计行公式
    month_task = month_task_data.get('revenue_budget', 0) if month_task_data else 0
    online_target = month_task_data.get('online_revenue_target', 0) if month_task_data else 0
    offline_target = month_task_data.get('offline_revenue_target', 0) if month_task_data else 0
    
    ws['C4'] = month_task  # 月目标合计
    ws['D4'] = '=E4/C4'  # 总目标完成率
    ws['E4'] = f'=SUM(E5:E{last_row})'  # 总收入
    ws['F4'] = f'=SUM(F5:F{last_row})'  # 客房收入
    ws['G4'] = '=H4*I4'  # 客房Revpar
    ws['H4'] = '=N4/O4'  # 出租率
    ws['I4'] = '=F4/N4'  # 平均房价
    ws['J4'] = f'=(N4-BG4)/O4'  # 过夜出租率
    ws['K4'] = f'=(F4-BF4)/(N4-BG4)'  # 过夜均价
    ws['L4'] = f'=(F4-BF4)/O4'  # 过夜Revpar
    ws['M4'] = '=S4/N4'  # OTA间夜占比
    ws['N4'] = f'=SUM(N5:N{last_row})'  # 售房数
    ws['O4'] = f'=SUM(O5:O{last_row})'  # 可售房数量
    
    # 线上合计
    ws['P4'] = online_target  # 线上目标
    ws['Q4'] = '=R4/P4' if online_target > 0 else '=0'  # 线上目标完成率
    ws['R4'] = f'=U4+W4+Y4+AC4+AE4+AG4'  # 总收入（不含booking）
    ws['S4'] = f'=V4+X4+Z4+AD4+AF4+AH4'  # 售房数
    ws['T4'] = '=R4/S4'  # 平均房价
    
    # 携程
    ws['U4'] = f'=SUM(U5:U{last_row})'
    ws['V4'] = f'=SUM(V5:V{last_row})'
    # 美团
    ws['W4'] = f'=SUM(W5:W{last_row})'
    ws['X4'] = f'=SUM(X5:X{last_row})'
    # 飞猪
    ws['Y4'] = f'=SUM(Y5:Y{last_row})'
    ws['Z4'] = f'=SUM(Z5:Z{last_row})'
    # booking
    ws['AA4'] = f'=SUM(AA5:AA{last_row})'
    ws['AB4'] = f'=SUM(AB5:AB{last_row})'
    # 抖音
    ws['AC4'] = f'=SUM(AC5:AC{last_row})'
    ws['AD4'] = f'=SUM(AD5:AD{last_row})'
    # 其他2
    ws['AE4'] = f'=SUM(AE5:AE{last_row})'
    ws['AF4'] = f'=SUM(AF5:AF{last_row})'
    # 其他3
    ws['AG4'] = f'=SUM(AG5:AG{last_row})'
    ws['AH4'] = f'=SUM(AH5:AH{last_row})'
    
    # 线下合计
    ws['AI4'] = offline_target  # 线下目标
    ws['AJ4'] = '=AK4/AI4' if offline_target > 0 else '=0'  # 线下目标完成率
    ws['AK4'] = f'=AN4+AP4+AR4+AT4+AV4+AX4+AZ4+BB4+BD4+BF4'  # 总收入
    ws['AL4'] = f'=AO4+AQ4+AS4+AU4+AW4+AY4+BA4+BC4+BE4+BG4'  # 售房数
    ws['AM4'] = '=AK4/AL4'  # 平均房价
    
    # 散客
    ws['AN4'] = f'=SUM(AN5:AN{last_row})'
    ws['AO4'] = f'=SUM(AO5:AO{last_row})'
    # 会员体验
    ws['AP4'] = f'=SUM(AP5:AP{last_row})'
    ws['AQ4'] = f'=SUM(AQ5:AQ{last_row})'
    # 网络体验
    ws['AR4'] = f'=SUM(AR5:AR{last_row})'
    ws['AS4'] = f'=SUM(AS5:AS{last_row})'
    # 团队
    ws['AT4'] = f'=SUM(AT5:AT{last_row})'
    ws['AU4'] = f'=SUM(AU5:AU{last_row})'
    # 协议客户
    ws['AV4'] = f'=SUM(AV5:AV{last_row})'
    ws['AW4'] = f'=SUM(AW5:AW{last_row})'
    # 微信
    ws['AX4'] = f'=SUM(AX5:AX{last_row})'
    ws['AY4'] = f'=SUM(AY5:AY{last_row})'
    # 免费房
    ws['AZ4'] = f'=SUM(AZ5:AZ{last_row})'
    ws['BA4'] = f'=SUM(BA5:BA{last_row})'
    # 集团金卡
    ws['BB4'] = f'=SUM(BB5:BB{last_row})'
    ws['BC4'] = f'=SUM(BC5:BC{last_row})'
    # 集团黑金卡
    ws['BD4'] = f'=SUM(BD5:BD{last_row})'
    ws['BE4'] = f'=SUM(BE5:BE{last_row})'
    # 钟点房
    ws['BF4'] = f'=SUM(BF5:BF{last_row})'
    ws['BG4'] = f'=SUM(BG5:BG{last_row})'
    # 其他收入合计
    ws['BH4'] = f'=SUM(BH5:BH{last_row})'
    ws['BI4'] = f'=SUM(BI5:BI{last_row})'
    ws['BJ4'] = f'=SUM(BJ5:BJ{last_row})'
    ws['BK4'] = f'=SUM(BK5:BK{last_row})'
    ws['BL4'] = f'=SUM(BL5:BL{last_row})'
    ws['BM4'] = f'=SUM(BM5:BM{last_row})'
    ws['BN4'] = f'=SUM(BN5:BN{last_row})'
    # 好评合计
    ws['BO4'] = '=BR4+BT4+BV4'
    ws['BP4'] = '=BS4+BU4+BW4'
    ws['BQ4'] = '=BP4/BO4'
    # 携程、艺龙、去哪儿
    ws['BR4'] = f'=SUM(BR5:BR{last_row})'
    ws['BS4'] = f'=SUM(BS5:BS{last_row})'
    # 美团
    ws['BT4'] = f'=SUM(BT5:BT{last_row})'
    ws['BU4'] = f'=SUM(BU5:BU{last_row})'
    # 飞猪
    ws['BV4'] = f'=SUM(BV5:BV{last_row})'
    ws['BW4'] = f'=SUM(BW5:BW{last_row})'
    # 携程流量
    ws['BX4'] = f'=SUM(BX5:BX{last_row})'
    ws['BY4'] = f'=SUM(BY5:BY{last_row})'
    ws['BZ4'] = f'=SUM(BZ5:BZ{last_row})'
    ws['CA4'] = f'=SUM(CA5:CA{last_row})'
    # 美团流量
    ws['CB4'] = f'=SUM(CB5:CB{last_row})'
    ws['CC4'] = f'=SUM(CC5:CC{last_row})'
    ws['CD4'] = f'=SUM(CD5:CD{last_row})'
    # 私域流量
    ws['CE4'] = f'=(CF4-BG4)/N4'
    ws['CF4'] = f'=SUM(CF5:CF{last_row})'
    ws['CG4'] = f'=SUM(CG5:CG{last_row})'
    # 私域订单
    ws['CH4'] = f'=SUM(CH5:CH{last_row})'
    ws['CI4'] = f'=SUM(CI5:CI{last_row})'
    # 会员
    ws['CJ4'] = f'=SUM(CJ5:CJ{last_row})'
    ws['CK4'] = f'=SUM(CK5:CK{last_row})'
    ws['CL4'] = f'=SUM(CL5:CL{last_row})'
    
    # ==================== 数据行 ====================
    for idx, report in enumerate(reports_data):
        row = 5 + idx
        data = report.get('data', {})
        
        # 日期
        report_date = report.get('report_date', '')
        ws.cell(row=row, column=1, value=report_date)
        
        # 星期
        if report_date:
            try:
                date_obj = datetime.strptime(report_date, '%Y-%m-%d')
                weekday = week_array[date_obj.weekday()]
                ws.cell(row=row, column=2, value=weekday)
                # 周五、周六标红
                if weekday in ['星期五', '星期六']:
                    ws.cell(row=row, column=1).fill = pink_fill
                    ws.cell(row=row, column=2).fill = pink_fill
            except:
                pass
        
        # 可售房数量
        salable_rooms = data.get('salable_rooms', 59)
        
        # C-O: 合计区域
        ws.cell(row=row, column=3, value=month_task)  # 月任务
        ws.cell(row=row, column=4, value=f'=E{row}/C{row}')  # 总目标完成率
        ws.cell(row=row, column=5, value=f'=R{row}+AK{row}+BH{row}')  # 总收入
        ws.cell(row=row, column=6, value=f'=R{row}+AK{row}')  # 客房收入
        ws.cell(row=row, column=7, value=f'=I{row}*H{row}')  # 客房Revpar
        ws.cell(row=row, column=8, value=f'=N{row}/O{row}')  # 出租率
        ws.cell(row=row, column=9, value=f'=F{row}/N{row}')  # 平均房价
        ws.cell(row=row, column=10, value=f'=(N{row}-BG{row})/O{row}')  # 过夜出租率
        ws.cell(row=row, column=11, value=f'=(F{row}-BF{row})/(N{row}-BG{row})')  # 过夜均价
        ws.cell(row=row, column=12, value=f'=(F{row}-BF{row})/O{row}')  # 过夜Revpar
        ws.cell(row=row, column=13, value=f'=S{row}/N{row}')  # OTA间夜占比
        ws.cell(row=row, column=14, value=f'=S{row}+AL{row}')  # 售房数
        ws.cell(row=row, column=15, value=salable_rooms)  # 可售房数量
        
        # P-T: 线上合计
        ws.cell(row=row, column=16, value=online_target)  # 线上目标
        ws.cell(row=row, column=17, value=f'=R{row}/P{row}' if online_target > 0 else '=0')  # 线上目标完成率
        # 线上总收入 = 携程+美团+飞猪+抖音+同程+去哪儿+智行
        ws.cell(row=row, column=18, value=f'=U{row}+W{row}+Y{row}+AC{row}+AE{row}+AG{row}')  # 总收入
        # 线上总间夜 = 携程+美团+飞猪+抖音+同程+去哪儿+智行
        ws.cell(row=row, column=19, value=f'=V{row}+X{row}+Z{row}+AD{row}+AF{row}+AH{row}')  # 售房数
        ws.cell(row=row, column=20, value=f'=R{row}/S{row}' if True else '=0')  # 平均房价
        
        # 携程 U-V
        ws.cell(row=row, column=21, value=data.get('xb_revenue', 0))  # 收入
        ws.cell(row=row, column=22, value=data.get('xb_rooms', 0))  # 间夜
        
        # 美团 W-X
        ws.cell(row=row, column=23, value=data.get('mt_revenue', 0))  # 收入
        ws.cell(row=row, column=24, value=data.get('mt_rooms', 0))  # 间夜
        
        # 飞猪 Y-Z
        ws.cell(row=row, column=25, value=data.get('fliggy_revenue', 0))  # 收入
        ws.cell(row=row, column=26, value=data.get('fliggy_rooms', 0))  # 间夜
        
        # booking AA-AB (数据库中没有booking字段，留空)
        ws.cell(row=row, column=27, value=0)  # 收入
        ws.cell(row=row, column=28, value=0)  # 间夜
        
        # 抖音 AC-AD
        ws.cell(row=row, column=29, value=data.get('dy_revenue', 0))  # 收入
        ws.cell(row=row, column=30, value=data.get('dy_rooms', 0))  # 间夜
        
        # 其他2 AE-AF (同程 tc)
        ws.cell(row=row, column=31, value=data.get('tc_revenue', 0))  # 收入
        ws.cell(row=row, column=32, value=data.get('tc_rooms', 0))  # 间夜
        
        # 其他3 AG-AH (去哪儿 qn + 智行 zx)
        qn_revenue = data.get('qn_revenue', 0) or 0
        zx_revenue = data.get('zx_revenue', 0) or 0
        qn_rooms = data.get('qn_rooms', 0) or 0
        zx_rooms = data.get('zx_rooms', 0) or 0
        ws.cell(row=row, column=33, value=qn_revenue + zx_revenue)  # 去哪儿+智行收入
        ws.cell(row=row, column=34, value=qn_rooms + zx_rooms)  # 去哪儿+智行间夜
        
        # AI-AM: 线下合计
        ws.cell(row=row, column=35, value=offline_target)  # 线下目标
        ws.cell(row=row, column=36, value=f'=AK{row}/AI{row}' if offline_target > 0 else '=0')  # 线下目标完成率
        # 线下总收入 = 散客+会员体验+网络体验+团队+协议+微信+免费+金卡+黑金+钟点
        ws.cell(row=row, column=37, value=f'=AN{row}+AP{row}+AR{row}+AT{row}+AV{row}+AX{row}+AZ{row}+BB{row}+BD{row}+BF{row}')  # 总收入
        # 线下总间夜
        ws.cell(row=row, column=38, value=f'=AO{row}+AQ{row}+AS{row}+AU{row}+AW{row}+AY{row}+BA{row}+BC{row}+BE{row}+BG{row}')  # 售房数
        ws.cell(row=row, column=39, value=f'=AK{row}/AL{row}' if True else '=0')  # 平均房价
        
        # 散客 AN-AO
        ws.cell(row=row, column=40, value=data.get('walkin_revenue', 0))  # 收入
        ws.cell(row=row, column=41, value=data.get('walkin_rooms', 0))  # 间夜
        
        # 会员体验 AP-AQ
        ws.cell(row=row, column=42, value=data.get('member_exp_revenue', 0))  # 收入
        ws.cell(row=row, column=43, value=data.get('member_exp_rooms', 0))  # 间夜
        
        # 网络体验 AR-AS
        ws.cell(row=row, column=44, value=data.get('web_exp_revenue', 0))  # 收入
        ws.cell(row=row, column=45, value=data.get('web_exp_rooms', 0))  # 间夜
        
        # 团队 AT-AU
        ws.cell(row=row, column=46, value=data.get('group_revenue', 0))  # 收入
        ws.cell(row=row, column=47, value=data.get('group_rooms', 0))  # 间夜
        
        # 协议客户 AV-AW
        ws.cell(row=row, column=48, value=data.get('protocol_revenue', 0))  # 收入
        ws.cell(row=row, column=49, value=data.get('protocol_rooms', 0))  # 间夜
        
        # 微信 AX-AY
        ws.cell(row=row, column=50, value=data.get('wechat_revenue', 0))  # 收入
        ws.cell(row=row, column=51, value=data.get('wechat_rooms', 0))  # 间夜
        
        # 免费房 AZ-BA
        ws.cell(row=row, column=52, value=data.get('free_revenue', 0))  # 收入
        ws.cell(row=row, column=53, value=data.get('free_rooms', 0))  # 间夜
        
        # 集团金卡 BB-BC
        ws.cell(row=row, column=54, value=data.get('gold_card_revenue', 0))  # 收入
        ws.cell(row=row, column=55, value=data.get('gold_card_rooms', 0))  # 间夜
        
        # 集团黑金卡 BD-BE
        ws.cell(row=row, column=56, value=data.get('black_gold_revenue', 0))  # 收入
        ws.cell(row=row, column=57, value=data.get('black_gold_rooms', 0))  # 间夜
        
        # 钟点房 BF-BG
        ws.cell(row=row, column=58, value=data.get('hourly_revenue', 0))  # 收入
        ws.cell(row=row, column=59, value=data.get('hourly_rooms', 0))  # 间夜
        
        # 其他收入合计 BH-BN
        ws.cell(row=row, column=60, value=f'=BI{row}+BJ{row}+BK{row}+BL{row}+BM{row}+BN{row}')  # 其他收入合计
        ws.cell(row=row, column=61, value=data.get('parking_revenue', 0))  # 停车费
        ws.cell(row=row, column=62, value=data.get('dining_revenue', 0))  # 餐饮
        ws.cell(row=row, column=63, value=data.get('meeting_revenue', 0))  # 会议活动
        ws.cell(row=row, column=64, value=data.get('goods_revenue', 0))  # 商品
        ws.cell(row=row, column=65, value=data.get('member_card_revenue', 0))  # 会员卡收入
        ws.cell(row=row, column=66, value=data.get('other_revenue', 0))  # 其他
        
        # 好评合计 BO-BQ
        ws.cell(row=row, column=67, value=f'=BR{row}+BT{row}+BV{row}')  # 订单数
        ws.cell(row=row, column=68, value=f'=BS{row}+BU{row}+BW{row}')  # 5分点评数
        ws.cell(row=row, column=69, value=f'=BP{row}/BO{row}')  # 好评转化率
        
        # 携程、艺龙、去哪儿 BR-BS
        ws.cell(row=row, column=70, value=data.get('xb_reviewable', 0))  # 可评价订单数
        ws.cell(row=row, column=71, value=data.get('xb_good_review', 0))  # 5分点评数量
        
        # 美团 BT-BU
        ws.cell(row=row, column=72, value=data.get('mt_reviewable', 0))  # 可评价订单数
        ws.cell(row=row, column=73, value=data.get('mt_good_review', 0))  # 5分点评数量
        
        # 飞猪 BV-BW
        ws.cell(row=row, column=74, value=data.get('fliggy_reviewable', 0))  # 可评价订单数
        ws.cell(row=row, column=75, value=data.get('fliggy_good_review', 0))  # 5分点评数量
        
        # 携程流量 BX-CA
        ws.cell(row=row, column=76, value=data.get('xb_exposure', 0))  # 列表页曝光量
        ws.cell(row=row, column=77, value=data.get('xb_exp_rate', 0))  # 曝光转化率
        ws.cell(row=row, column=78, value=data.get('xb_bk_rate', 0))  # 下单转化率
        ws.cell(row=row, column=79, value=data.get('xb_clinch_rate', 0))  # 成交转化率
        
        # 美团流量 CB-CD
        ws.cell(row=row, column=80, value=data.get('mt_exposure', 0))  # 曝光人数
        ws.cell(row=row, column=81, value=data.get('mt_click_rate', 0))  # 点击率
        ws.cell(row=row, column=82, value=data.get('mt_pay_rate', 0))  # 支付转化率
        
        # 私域流量 CE-CG
        ws.cell(row=row, column=83, value=f'=(CF{row}-BG{row})/N{row}')  # 微信加粉率
        ws.cell(row=row, column=84, value=data.get('wechat_add', 0))  # 微信加粉人数
        ws.cell(row=row, column=85, value=data.get('member_add', 0))  # 新增会员人数
        
        # 私域订单 CH-CI
        ws.cell(row=row, column=86, value=data.get('private_revenue', 0))  # 收入
        ws.cell(row=row, column=87, value=data.get('private_rooms', 0))  # 间夜
        
        # 会员 CJ-CL
        ws.cell(row=row, column=88, value=data.get('member_card_sold', 0))  # 售卡量
        ws.cell(row=row, column=89, value=data.get('member_card_revenue', 0))  # 售卡收入
        ws.cell(row=row, column=90, value=data.get('stored_value', 0))  # 储值
        
        # 应用样式
        for col in range(1, 91):
            cell = ws.cell(row=row, column=col)
            cell.alignment = center_align
            cell.border = thin_border
        
        ws.row_dimensions[row].height = 25
        
        # 颜色
        # 蓝色区域
        for col in range(3, 15):  # C-N
            ws.cell(row=row, column=col).fill = blue_fill
        for col in range(16, 21):  # P-T
            ws.cell(row=row, column=col).fill = blue_fill
        for col in range(35, 40):  # AI-AM
            ws.cell(row=row, column=col).fill = blue_fill
        ws.cell(row=row, column=60).fill = blue_fill  # BH
        for col in range(67, 70):  # BO-BQ
            ws.cell(row=row, column=col).fill = blue_fill
        ws.cell(row=row, column=83).fill = blue_fill  # CE
        
        # 绿色区域
        ws.cell(row=row, column=15).fill = green_fill  # O
        for col in range(21, 35):  # U-AH
            ws.cell(row=row, column=col).fill = green_fill
        for col in range(40, 60):  # AN-BG
            ws.cell(row=row, column=col).fill = green_fill
        for col in range(61, 67):  # BI-BN
            ws.cell(row=row, column=col).fill = green_fill
        for col in range(70, 83):  # BR-CD
            ws.cell(row=row, column=col).fill = green_fill
        for col in range(84, 91):  # CF-CL
            ws.cell(row=row, column=col).fill = green_fill
        
        # 数字格式
        # 百分比格式
        for col in [4, 8, 10, 13, 17, 36, 69, 83]:  # D, H, J, M, Q, AJ, BQ, CE
            ws.cell(row=row, column=col).number_format = '0.00%'
        
        # 数值格式
        for col in [7, 9, 12, 20, 39]:  # G, I, L, T, AM
            ws.cell(row=row, column=col).number_format = '0.00'
    
    # 合计行格式
    for col in [4, 8, 10, 13, 17, 36, 69, 83]:
        ws.cell(row=4, column=col).number_format = '0.00%'
    for col in [7, 9, 12, 20, 39]:
        ws.cell(row=4, column=col).number_format = '0.00'
    
    # 设置列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    
    # 输出到bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Usage: python export_daily_report.py <json_data>"}))
        sys.exit(1)
    
    try:
        input_data = json.loads(sys.argv[1])
        
        # 支持两种模式：single 和 batch
        mode = input_data.get('mode', 'batch')
        
        if mode == 'single':
            # 单条导出模式
            single_report = input_data.get('report', {})
            month_task = input_data.get('month_task', {})
            
            # 将单条报表转换为列表格式
            reports = [{
                'hotel_name': single_report.get('hotel_name', ''),
                'report_date': single_report.get('report_date', ''),
                'data': single_report.get('data', {}),
            }]
        else:
            # 批量导出模式
            reports = input_data.get('reports', [])
            month_task = input_data.get('month_task', {})
        
        excel_bytes = create_excel_report(reports, month_task)
        
        # 输出base64编码的Excel内容
        excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
        print(json.dumps({"success": True, "data": excel_base64}))
        
    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)
