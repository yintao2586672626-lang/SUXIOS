#!/usr/bin/env python3
"""Build a traceable local/Obsidian knowledge library from a user document folder.

The script never executes archived code. It extracts document text, records exact
source fingerprints, deduplicates byte-identical files, writes Obsidian notes,
and emits a compact SUXIOS method pack plus a top-level source manifest.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import io
import json
import os
import re
import subprocess
import sys
import tempfile
import time
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable


TEXT_EXTENSIONS = {"", ".txt", ".md", ".markdown", ".csv", ".json", ".toml", ".log"}
HTML_EXTENSIONS = {".html", ".htm"}
DOCUMENT_EXTENSIONS = TEXT_EXTENSIONS | HTML_EXTENSIONS | {
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx", ".emm", ".emmx"
}
ARCHIVE_EXTENSIONS = {".zip", ".rar"}
MAX_MEMBER_BYTES = 64 * 1024 * 1024
MAX_TEXT_CHARS = 2_000_000
MAX_ARCHIVE_MEMBERS = 4000
SEED_VERSION = "2026-08-14.1"


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_text(value: str) -> str:
    value = value.replace("\r\n", "\n").replace("\r", "\n").replace("\x00", "")
    value = re.sub(r"[ \t]+\n", "\n", value)
    value = re.sub(r"\n{4,}", "\n\n\n", value)
    return value.strip()


def decode_text(data: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "utf-16", "gb18030", "big5"):
        try:
            return data.decode(encoding), encoding
        except (UnicodeDecodeError, LookupError):
            continue
    return data.decode("utf-8", errors="replace"), "utf-8-replacement"


def redact_sensitive(text: str) -> tuple[str, int]:
    patterns = [
        re.compile(r"(?i)\b(authorization)\s*[:=]\s*(bearer\s+)?[^\s,;]{6,}"),
        re.compile(r"(?i)\b(cookie|set-cookie)\s*[:=]\s*[^\n]{6,}"),
        re.compile(
            r"(?i)\b(password|passwd|pwd|secret|api[_-]?key|access[_-]?token|refresh[_-]?token)"
            r"\s*[:=]\s*[^\s,;]{4,}"
        ),
    ]
    count = 0
    for pattern in patterns:
        text, replacements = pattern.subn(lambda match: match.group(1) + "=[REDACTED]", text)
        count += replacements
    return text, count


def html_to_text(raw: str) -> str:
    raw = re.sub(r"(?is)<(script|style|noscript|template)\b[^>]*>.*?</\1\s*>", "", raw)
    raw = re.sub(r"(?i)<br\s*/?>|</(?:p|div|li|tr|h[1-6])\s*>", "\n", raw)
    raw = re.sub(r"(?i)<li\b[^>]*>", "\n- ", raw)
    return html.unescape(re.sub(r"(?s)<[^>]+>", " ", raw))


def legacy_binary_text(data: bytes) -> str:
    """Best-effort text recovery for legacy DOC/XLS/PPT without claiming structure."""
    candidates: list[str] = []
    for match in re.finditer(rb"(?:[\x20-\x7e\x80-\xff]\x00){4,}", data):
        text = match.group(0).decode("utf-16le", errors="ignore").strip()
        if text:
            candidates.append(text)
    for match in re.finditer(rb"[\x20-\x7e\x80-\xff]{8,}", data):
        text = match.group(0).decode("gb18030", errors="ignore").strip()
        if text and sum(char.isprintable() for char in text) / max(1, len(text)) > 0.8:
            candidates.append(text)
    unique: dict[str, None] = {}
    for item in candidates:
        item = re.sub(r"\s+", " ", item).strip()
        if len(item) >= 4:
            unique[item] = None
    return "\n".join(unique.keys())


def extract_pdf(data: bytes) -> tuple[str, dict[str, Any]]:
    worker = Path(__file__).with_name("extract_pdf_text_worker.py")
    with tempfile.TemporaryDirectory(prefix="suxios-pdf-") as temporary:
        temporary_path = Path(temporary)
        input_path = temporary_path / "input.pdf"
        text_path = temporary_path / "text.txt"
        metadata_path = temporary_path / "metadata.json"
        input_path.write_bytes(data)
        try:
            result = subprocess.run(
                [sys.executable, str(worker), str(input_path), str(text_path), str(metadata_path)],
                capture_output=True,
                timeout=30,
                check=False,
            )
        except subprocess.TimeoutExpired as exc:
            raise TimeoutError("pdf_extraction_timeout_30s") from exc
        if result.returncode != 0:
            raise ValueError("pdf_worker_failed")
        text = text_path.read_text(encoding="utf-8") if text_path.exists() else ""
        metadata = json.loads(metadata_path.read_text(encoding="utf-8")) if metadata_path.exists() else {}
        metadata["isolated_worker"] = True
        metadata["timeout_seconds"] = 30
        return text, metadata


def xml_paragraph_text(xml_bytes: bytes, text_tag: str, paragraph_tag: str) -> list[str]:
    root = ET.fromstring(xml_bytes)
    lines: list[str] = []
    for paragraph in root.iter():
        if paragraph.tag.rsplit("}", 1)[-1] != paragraph_tag:
            continue
        values = [
            node.text or ""
            for node in paragraph.iter()
            if node.tag.rsplit("}", 1)[-1] == text_tag and (node.text or "").strip()
        ]
        text = "".join(values).strip()
        if text:
            lines.append(text)
    return lines


def extract_docx(data: bytes) -> tuple[str, dict[str, Any]]:
    lines: list[str] = []
    part_count = 0
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        names = [
            name for name in archive.namelist()
            if name == "word/document.xml"
            or re.match(r"^word/(?:header|footer)\d+\.xml$", name)
        ]
        for name in sorted(names, key=lambda item: (item != "word/document.xml", item)):
            raw = archive.read(name)
            if len(raw) > 16 * 1024 * 1024:
                raise ValueError("docx_xml_part_too_large")
            part_count += 1
            prefix = ""
            if "/header" in name:
                prefix = "[页眉] "
            elif "/footer" in name:
                prefix = "[页脚] "
            lines.extend(prefix + line for line in xml_paragraph_text(raw, "t", "p"))
    return "\n".join(lines), {"xml_part_count": part_count, "media_loaded": False}


def extract_pptx(data: bytes) -> tuple[str, dict[str, Any]]:
    lines: list[str] = []
    slide_count = 0
    note_count = 0
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        slide_names = [
            name for name in archive.namelist()
            if re.match(r"^ppt/slides/slide\d+\.xml$", name)
        ]
        slide_names.sort(key=lambda item: int(re.search(r"(\d+)\.xml$", item).group(1)))
        for index, name in enumerate(slide_names, 1):
            raw = archive.read(name)
            if len(raw) > 16 * 1024 * 1024:
                raise ValueError("pptx_slide_xml_too_large")
            slide_lines = xml_paragraph_text(raw, "t", "p")
            if slide_lines:
                lines.append(f"## 第 {index} 页\n" + "\n".join(slide_lines))
            slide_count += 1
        note_names = sorted(
            name for name in archive.namelist()
            if re.match(r"^ppt/notesSlides/notesSlide\d+\.xml$", name)
        )
        for name in note_names:
            raw = archive.read(name)
            if len(raw) > 8 * 1024 * 1024:
                continue
            note_lines = xml_paragraph_text(raw, "t", "p")
            if note_lines:
                note_count += 1
                lines.append("## 演讲者备注\n" + "\n".join(note_lines))
    return "\n\n".join(lines), {"slide_count": slide_count, "notes_part_count": note_count, "media_loaded": False}


def extract_xlsx(data: bytes) -> tuple[str, dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=False)
    lines: list[str] = []
    sheet_meta: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            sheet_lines: list[str] = []
            non_empty = 0
            truncated = False
            for row_number, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                values = []
                for column_number, value in enumerate(row, 1):
                    if value is None or value == "":
                        continue
                    non_empty += 1
                    if non_empty > 200_000:
                        truncated = True
                        break
                    values.append(f"R{row_number}C{column_number}={value}")
                if values:
                    sheet_lines.append(" | ".join(values))
                if truncated:
                    break
            lines.append(f"## 工作表：{worksheet.title}\n" + "\n".join(sheet_lines))
            sheet_meta.append({
                "name": worksheet.title,
                "non_empty_cells": min(non_empty, 200_000),
                "truncated": truncated,
            })
    finally:
        workbook.close()
    return "\n\n".join(lines), {"sheets": sheet_meta}


def extract_emmx(data: bytes) -> tuple[str, dict[str, Any]]:
    if zipfile.is_zipfile(io.BytesIO(data)):
        values: list[str] = []
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            for name in archive.namelist()[:200]:
                if name.lower().endswith((".xml", ".json", ".txt")):
                    raw = archive.read(name)
                    decoded, _ = decode_text(raw)
                    values.append(html_to_text(decoded) if "<" in decoded else decoded)
        return "\n".join(values), {"container": "zip"}
    return legacy_binary_text(data), {"container": "binary"}


def extract_document(data: bytes, extension: str) -> tuple[str, str, dict[str, Any]]:
    metadata: dict[str, Any] = {}
    if extension in TEXT_EXTENSIONS or extension == ".py":
        text, encoding = decode_text(data)
        metadata["encoding"] = encoding
        status = "extracted"
    elif extension in HTML_EXTENSIONS:
        raw, encoding = decode_text(data)
        text = html_to_text(raw)
        metadata["encoding"] = encoding
        status = "extracted"
    elif extension == ".pdf":
        text, metadata = extract_pdf(data)
        status = "extracted" if text.strip() else "ocr_required"
    elif extension == ".docx":
        text, metadata = extract_docx(data)
        status = "extracted"
    elif extension == ".pptx":
        text, metadata = extract_pptx(data)
        status = "extracted"
    elif extension == ".xlsx":
        try:
            text, metadata = extract_xlsx(data)
            status = "extracted"
        except zipfile.BadZipFile:
            if data.startswith(b"\xd0\xcf\x11\xe0"):
                text = legacy_binary_text(data)
                metadata = {"detected_container": "ole2", "declared_extension": "xlsx"}
                status = "degraded_mislabeled_legacy_excel" if text.strip() else "converter_required"
            else:
                raise
    elif extension in {".doc", ".xls", ".ppt"}:
        text = legacy_binary_text(data)
        status = "degraded_legacy_binary_extraction" if text.strip() else "converter_required"
    elif extension in {".emm", ".emmx"}:
        text, metadata = extract_emmx(data)
        status = "extracted" if text.strip() else "converter_required"
    else:
        return "", "unsupported", metadata

    text = normalize_text(text)
    if len(text) > MAX_TEXT_CHARS:
        text = text[:MAX_TEXT_CHARS] + "\n\n[正文超过 2,000,000 字，已在知识副本中截断；原文件与哈希仍保留。]"
        metadata["text_truncated"] = True
    text, redaction_count = redact_sensitive(text)
    if redaction_count:
        metadata["sensitive_value_redactions"] = redaction_count
    return text, status, metadata


def safe_archive_member(name: str) -> bool:
    normalized = name.replace("\\", "/")
    return bool(normalized) and not normalized.startswith("/") and "../" not in f"/{normalized}"


def iter_zip_documents(path: Path) -> Iterable[tuple[str, bytes, dict[str, Any]]]:
    with zipfile.ZipFile(path) as archive:
        infos = archive.infolist()
        if len(infos) > MAX_ARCHIVE_MEMBERS:
            return
        for info in infos:
            if info.is_dir() or not safe_archive_member(info.filename):
                continue
            extension = Path(info.filename).suffix.lower()
            if extension not in DOCUMENT_EXTENSIONS or info.file_size > MAX_MEMBER_BYTES:
                continue
            if info.flag_bits & 0x1:
                yield info.filename, b"", {"archive_member_status": "encrypted_skipped"}
                continue
            try:
                yield info.filename, archive.read(info), {"archive_member_size": info.file_size}
            except Exception as exc:  # noqa: BLE001
                yield info.filename, b"", {"archive_member_status": f"read_failed:{type(exc).__name__}"}


def iter_rar_documents(path: Path) -> Iterable[tuple[str, bytes, dict[str, Any]]]:
    try:
        listing = subprocess.run(
            ["tar", "-tf", str(path)], capture_output=True, timeout=30, check=False
        )
    except (OSError, subprocess.TimeoutExpired):
        return
    names, _ = decode_text(listing.stdout)
    members = [line.strip() for line in names.splitlines() if line.strip()]
    if listing.returncode != 0 or len(members) > MAX_ARCHIVE_MEMBERS:
        return
    if any(not safe_archive_member(member) for member in members):
        return
    document_members = [member for member in members if Path(member).suffix.lower() in DOCUMENT_EXTENSIONS]
    with tempfile.TemporaryDirectory(prefix="suxios-rar-") as temporary:
        try:
            result = subprocess.run(
                ["tar", "-xf", str(path), "-C", temporary], capture_output=True, timeout=90, check=False
            )
        except (OSError, subprocess.TimeoutExpired):
            return
        if result.returncode != 0:
            return
        extraction_root = Path(temporary).resolve()
        for member in document_members:
            member_path = (extraction_root / member).resolve()
            try:
                member_path.relative_to(extraction_root)
            except ValueError:
                yield member, b"", {"archive_member_status": "path_escape_skipped"}
                continue
            if not member_path.is_file() or member_path.stat().st_size > MAX_MEMBER_BYTES:
                yield member, b"", {"archive_member_status": "missing_or_too_large"}
                continue
            try:
                data = member_path.read_bytes()
            except OSError:
                yield member, b"", {"archive_member_status": "read_failed"}
                continue
            yield member, data, {"archive_member_size": len(data), "archive_extraction": "single_bounded_temp_pass"}


def classify(title: str, text: str = "") -> tuple[str, str]:
    sample = f"{title}\n{text[:4000]}".lower()
    if "skill.md" in sample or "agents/openai" in sample or "prompt" in sample or "提示词" in sample:
        classification = "untrusted_skill_or_prompt_reference"
    elif any(term in sample for term in ("合同", "订单", "名单", "报价单", "个人", "手机号")):
        classification = "hotel_specific_or_sensitive_reference"
    elif title.startswith(("方法_", "专题_")) or "sop" in sample:
        classification = "industry_method_reference"
    elif "案例" in title or "报告" in title:
        classification = "case_reference"
    else:
        classification = "general_reference"

    topic_rules = [
        ("预订进度", ("预订进度", "booking pace", "提前期", "pickup")),
        ("收益管理", ("收益", "revpar", "adr", "occ", "出租率")),
        ("OTA运营", ("ota", "携程", "美团", "点评", "流量")),
        ("定价与房态", ("定价", "调价", "价格", "房态", "库存")),
        ("竞争商圈", ("竞争圈", "商圈", "竞品", "竞争态势")),
        ("房型产品", ("房型", "产品", "早餐", "设施")),
        ("口碑服务", ("口碑", "差评", "满意度", "服务质量")),
        ("渠道与订单", ("渠道", "订单", "分销")),
        ("投资与商业", ("投资", "回报", "商业计划", "预算", "测算")),
        ("培训与SOP", ("培训", "sop", "操作手册", "流程", "话术")),
        ("AI与系统", ("ai", "智能", "系统", "数字化", "skill")),
    ]
    for topic, terms in topic_rules:
        if any(term in sample for term in terms):
            return classification, topic
    return classification, "其他资料"


def note_slug(title: str, digest: str) -> str:
    slug = re.sub(r"[<>:\"/\\|?*\x00-\x1f]", "_", title).strip(" ._")
    slug = re.sub(r"\s+", " ", slug)
    if not slug:
        slug = "未命名资料"
    return f"{slug[:82]}__{digest[:10]}"


def yaml_value(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False)


def write_utf8(path: Path, value: str) -> None:
    """Write deterministic UTF-8 bytes without Windows newline translation."""
    path.write_bytes(value.encode("utf-8"))


def write_note(path: Path, record: dict[str, Any], text: str) -> None:
    sources = record.get("source_paths", [record["source_path"]])
    frontmatter = [
        "---",
        f"title: {yaml_value(record['title'])}",
        "type: ai-knowledge-source",
        f"topic: {yaml_value(record['topic'])}",
        f"classification: {yaml_value(record['classification'])}",
        f"extraction_status: {yaml_value(record['extraction_status'])}",
        f"source_sha256: {record['sha256']}",
        f"text_sha256: {record.get('text_sha256', '')}",
        f"source_paths: {yaml_value(sources)}",
        "decision_safe: false",
        "external_write_authorized: false",
        f"imported_at: {yaml_value(record['imported_at'])}",
        "---",
        "",
        f"# {record['title']}",
        "",
        "> 这是来源可追溯的参考资料，不自动等于当前酒店、当前平台或当前业务日事实；执行前必须回到真实数据与人工复核。",
        "",
        "## 来源",
        "",
    ]
    for source in sources:
        frontmatter.append(f"- `{source}`")
    frontmatter.extend([
        "",
        "## 解析状态",
        "",
        f"- 状态：`{record['extraction_status']}`",
        f"- 原文件 SHA-256：`{record['sha256']}`",
        f"- 提取字符数：{record.get('char_count', 0)}",
    ])
    metadata = record.get("extraction_metadata") or {}
    if metadata:
        frontmatter.extend(["- 解析元数据：", "", "```json", json.dumps(metadata, ensure_ascii=False, indent=2), "```"])
    frontmatter.extend(["", "## 提取正文", "", text if text else "_未取得可检索正文；请查看解析状态与原文件。_", ""])
    write_utf8(path, "\n".join(frontmatter))


def curated_source_name(record: dict[str, Any]) -> str:
    if record.get("archive_member") or record.get("extension") != ".md":
        return ""
    for source_path in record.get("source_paths", [record.get("source_path", "")]):
        name = Path(source_path).name
        if "(1)" in name:
            continue
        if re.match(r"^(方法_|专题_|案例_|方法库\.md$|0\d_.*案例|1[0-3]_.*案例|99_30个案例)", name):
            return name
    return ""


def is_curated_markdown(record: dict[str, Any]) -> bool:
    return curated_source_name(record) != ""


def markdown_title(filename: str, text: str) -> str:
    match = re.search(r"(?m)^#\s+(.+?)\s*$", text)
    return match.group(1).strip() if match else Path(filename).stem


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", required=True)
    parser.add_argument("--obsidian", required=True)
    parser.add_argument("--suxios-output", required=True)
    parser.add_argument("--replace", action="store_true")
    parser.add_argument("--progress-log")
    args = parser.parse_args()

    source = Path(args.source).resolve()
    obsidian = Path(args.obsidian).resolve()
    suxios_output = Path(args.suxios_output).resolve()
    if not source.is_dir():
        raise SystemExit("source_directory_missing")
    if not obsidian.parent.exists():
        raise SystemExit("obsidian_parent_missing")
    if obsidian.exists() and not args.replace:
        raise SystemExit("obsidian_output_exists_use_replace")

    imported_at = time.strftime("%Y-%m-%dT%H:%M:%S%z")
    override_path = suxios_output / "ocr-overrides.json"
    ocr_overrides: dict[str, Any] = {}
    if override_path.is_file():
        decoded_overrides = json.loads(override_path.read_text(encoding="utf-8"))
        ocr_overrides = decoded_overrides.get("entries", {}) if isinstance(decoded_overrides, dict) else {}
    top_files = sorted((path for path in source.rglob("*") if path.is_file()), key=lambda item: str(item).lower())
    records: list[dict[str, Any]] = []
    extracted_payloads: dict[str, str] = {}
    archive_summaries: list[dict[str, Any]] = []

    def process_bytes(
        data: bytes,
        display_path: str,
        title: str,
        top_level: bool,
        archive_path: str = "",
        archive_member: str = "",
        extra: dict[str, Any] | None = None,
    ) -> None:
        extension = Path(title).suffix.lower()
        digest = sha256_bytes(data)
        metadata = dict(extra or {})
        try:
            text, status, extraction_metadata = extract_document(data, extension)
            metadata.update(extraction_metadata)
        except Exception as exc:  # noqa: BLE001
            text = ""
            status = f"extract_failed:{type(exc).__name__}"
            metadata["failure_code"] = re.sub(r"[^a-zA-Z0-9:_-]+", "_", str(exc))[:160]
        override = ocr_overrides.get(digest)
        if status == "ocr_required" and isinstance(override, dict):
            override_text = normalize_text(str(override.get("text", "")))
            if override_text:
                text = override_text
                status = "extracted_visual_review"
                metadata.update({
                    "visual_review_page_count": int(override.get("page_count", 0)),
                    "visual_review_method": "embedded_page_image_visual_review",
                    "visual_reviewed_at": "2026-08-15",
                })
        classification, topic = classify(title, text)
        record = {
            "title": title,
            "source_path": display_path,
            "source_paths": [display_path],
            "top_level": top_level,
            "archive_path": archive_path,
            "archive_member": archive_member,
            "extension": extension or "[no-extension]",
            "bytes": len(data),
            "sha256": digest,
            "text_sha256": sha256_bytes(text.encode("utf-8")) if text else "",
            "char_count": len(text),
            "extraction_status": status,
            "extraction_metadata": metadata,
            "classification": classification,
            "topic": topic,
            "decision_safe": False,
            "external_write_authorized": False,
            "imported_at": imported_at,
        }
        records.append(record)
        extracted_payloads.setdefault(digest, text)

    for index, path in enumerate(top_files, 1):
        relative = path.relative_to(source).as_posix()
        extension = path.suffix.lower()
        if args.progress_log:
            Path(args.progress_log).parent.mkdir(parents=True, exist_ok=True)
            with Path(args.progress_log).open("a", encoding="utf-8") as progress_handle:
                progress_handle.write(json.dumps({"progress": index, "total": len(top_files), "file": relative}, ensure_ascii=False) + "\n")
        if index == 1 or index % 10 == 0 or index == len(top_files):
            print(json.dumps({"progress": index, "total": len(top_files), "file": relative}, ensure_ascii=False), flush=True)
        try:
            data = path.read_bytes()
        except OSError as exc:
            records.append({
                "title": path.name,
                "source_path": str(path),
                "source_paths": [str(path)],
                "top_level": True,
                "extension": extension or "[no-extension]",
                "bytes": path.stat().st_size if path.exists() else 0,
                "sha256": "",
                "text_sha256": "",
                "char_count": 0,
                "extraction_status": f"read_failed:{type(exc).__name__}",
                "extraction_metadata": {},
                "classification": "general_reference",
                "topic": "其他资料",
                "decision_safe": False,
                "external_write_authorized": False,
                "imported_at": imported_at,
            })
            continue

        if extension in ARCHIVE_EXTENSIONS:
            archive_digest = sha256_bytes(data)
            members: list[tuple[str, bytes, dict[str, Any]]] = []
            try:
                members = list(iter_zip_documents(path) if extension == ".zip" else iter_rar_documents(path))
                archive_status = "inventory_and_document_members_extracted" if members else "inventory_only_no_supported_members"
            except Exception as exc:  # noqa: BLE001
                archive_status = f"archive_inventory_failed:{type(exc).__name__}"
            archive_summaries.append({
                "path": str(path),
                "sha256": archive_digest,
                "bytes": len(data),
                "status": archive_status,
                "document_member_count": len(members),
                "code_execution": "not_executed",
            })
            records.append({
                "title": path.name,
                "source_path": str(path),
                "source_paths": [str(path)],
                "top_level": True,
                "archive_path": "",
                "archive_member": "",
                "extension": extension,
                "bytes": len(data),
                "sha256": archive_digest,
                "text_sha256": "",
                "char_count": 0,
                "extraction_status": archive_status,
                "extraction_metadata": {"document_member_count": len(members), "code_execution": "not_executed"},
                "classification": "untrusted_archive_reference",
                "topic": "AI与系统",
                "decision_safe": False,
                "external_write_authorized": False,
                "imported_at": imported_at,
            })
            for member, member_data, extra in members:
                process_bytes(
                    member_data,
                    f"{path}!{member}",
                    Path(member).name,
                    False,
                    str(path),
                    member,
                    extra,
                )
            continue

        process_bytes(data, str(path), path.name, True)

    # Merge exact duplicate payloads while preserving every source path.
    canonical_by_sha: dict[str, dict[str, Any]] = {}
    for record in records:
        digest = record.get("sha256") or f"missing:{record['source_path']}"
        existing = canonical_by_sha.get(digest)
        if existing is None:
            canonical_by_sha[digest] = record
            continue
        existing["source_paths"].extend(record.get("source_paths", []))
        existing["source_paths"] = sorted(set(existing["source_paths"]))
        existing["duplicate_count"] = len(existing["source_paths"]) - 1
        record["canonical_sha256"] = digest
        record["duplicate_of"] = existing["source_path"]

    canonical_records = list(canonical_by_sha.values())
    if obsidian.exists():
        import shutil
        # Replace only files owned by this generator. User-curated folders such
        # as 03-重点吸收 must survive a full source refresh.
        for generated_dir in ("02-主题索引", "03-方法库", "04-资料正文"):
            target = obsidian / generated_dir
            if target.is_dir():
                shutil.rmtree(target)
        for generated_file in ("00_导入说明.md", "01_资料总索引.md", "_manifest.json", "_manifest.csv"):
            target = obsidian / generated_file
            if target.is_file():
                target.unlink()
    notes_dir = obsidian / "04-资料正文"
    topics_dir = obsidian / "02-主题索引"
    methods_dir = obsidian / "03-方法库"
    notes_dir.mkdir(parents=True, exist_ok=True)
    topics_dir.mkdir(parents=True, exist_ok=True)
    methods_dir.mkdir(parents=True, exist_ok=True)
    suxios_output.mkdir(parents=True, exist_ok=True)

    topic_links: dict[str, list[str]] = defaultdict(list)
    note_map: dict[str, str] = {}
    for record in sorted(canonical_records, key=lambda item: (item["topic"], item["title"].lower())):
        if record["extension"] in ARCHIVE_EXTENSIONS:
            continue
        digest = record.get("sha256") or sha256_bytes(record["source_path"].encode("utf-8"))
        slug = note_slug(record["title"], digest)
        target_dir = methods_dir if is_curated_markdown(record) else notes_dir
        target = target_dir / f"{slug}.md"
        write_note(target, record, extracted_payloads.get(digest, ""))
        relative_note = target.relative_to(obsidian).with_suffix("").as_posix()
        note_map[digest] = relative_note
        topic_links[record["topic"]].append(f"- [[{relative_note}|{Path(record['title']).stem}]]")

    for topic, links in sorted(topic_links.items()):
        path = topics_dir / f"{note_slug(topic, sha256_bytes(topic.encode('utf-8')))[:90]}.md"
        write_utf8(path,
            "\n".join([
                "---", f"title: {yaml_value(topic)}", "type: ai-knowledge-topic-index", "---", "",
                f"# {topic}", "", *sorted(set(links)), "",
            ]),
        )

    top_level_records = [record for record in records if record.get("top_level")]
    unique_top_level_sha = {record.get("sha256") for record in top_level_records if record.get("sha256")}
    status_counts = Counter(record["extraction_status"] for record in records)
    extension_counts = Counter(record["extension"] for record in top_level_records)
    topic_counts = Counter(record["topic"] for record in canonical_records if record["extension"] not in ARCHIVE_EXTENSIONS)
    summary = {
        "schema_version": 1,
        "seed_version": SEED_VERSION,
        "source_root": str(source),
        "imported_at": imported_at,
        "top_level_file_count": len(top_level_records),
        "top_level_unique_sha256_count": len(unique_top_level_sha),
        "archive_document_member_count": sum(1 for record in records if record.get("archive_member")),
        "canonical_document_count": sum(1 for record in canonical_records if record["extension"] not in ARCHIVE_EXTENSIONS),
        "duplicate_source_count": len(records) - len(canonical_records),
        "extraction_status_counts": dict(sorted(status_counts.items())),
        "top_level_extension_counts": dict(sorted(extension_counts.items())),
        "topic_counts": dict(sorted(topic_counts.items())),
        "decision_boundary": "reference_only_not_current_hotel_fact",
        "external_write_authorized": False,
    }

    manifest_records = []
    for record in top_level_records:
        manifest_records.append({
            key: record.get(key)
            for key in (
                "title", "source_path", "extension", "bytes", "sha256", "text_sha256", "char_count",
                "extraction_status", "extraction_metadata", "classification", "topic", "duplicate_of", "decision_safe",
                "external_write_authorized",
            )
            if record.get(key) not in (None, "")
        })
    manifest = {"summary": summary, "records": manifest_records, "archives": archive_summaries}
    manifest_json = json.dumps(manifest, ensure_ascii=False, indent=2)
    write_utf8(obsidian / "_manifest.json", manifest_json)
    write_utf8(suxios_output / "source-manifest.json", manifest_json)

    with (obsidian / "_manifest.csv").open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=[
            "title", "source_path", "extension", "bytes", "sha256", "text_sha256", "char_count",
            "extraction_status", "classification", "topic", "duplicate_of",
        ])
        writer.writeheader()
        for record in manifest_records:
            writer.writerow({key: record.get(key, "") for key in writer.fieldnames or []})

    method_entries_by_digest: dict[str, dict[str, Any]] = {}
    for record in top_level_records:
        source_name = curated_source_name(record)
        if source_name == "":
            continue
        text = extracted_payloads.get(record["sha256"], "")
        method_entries_by_digest[record["sha256"]] = {
            "key": "ai_library:" + record["sha256"][:16],
            "title": markdown_title(source_name, text),
            "source_filename": source_name,
            "source_path": record["source_path"],
            "source_sha256": record["sha256"],
            "text_sha256": record["text_sha256"],
            "topic": record["topic"],
            "classification": record["classification"],
            "body_markdown": text,
            "evidence_grade": "D",
            "decision_policy": "reference_only_human_review_required",
            "decision_safe": False,
            "external_write_authorized": False,
        }
    method_entries = list(method_entries_by_digest.values())
    method_pack = {
        "schema_version": 1,
        "seed_version": SEED_VERSION,
        "source_manifest_sha256": sha256_bytes(manifest_json.encode("utf-8")),
        "source_root": str(source),
        "imported_at": imported_at,
        "entry_count": len(method_entries),
        "entries": sorted(method_entries, key=lambda item: item["title"]),
        "boundary": {
            "scope": "industry_general_and_case_reference",
            "contains_current_hotel_fact": False,
            "decision_safe": False,
            "task_draft_safe": False,
            "external_write_authorized": False,
            "required_before_business_use": [
                "system_hotel_identity", "platform_or_pms_source", "business_date",
                "fresh_collection_or_verified_import", "save_readback", "human_review",
            ],
        },
    }
    method_pack_json = json.dumps(method_pack, ensure_ascii=False, indent=2)
    write_utf8(suxios_output / "method-pack.json", method_pack_json)

    topic_index_links = [
        f"- [[02-主题索引/{path.stem}|{topic}]]（{topic_counts[topic]}）"
        for topic, path in sorted((topic, next(topics_dir.glob(f"*__{sha256_bytes(topic.encode('utf-8'))[:10]}.md"))) for topic in topic_counts)
    ]
    summary_lines = [
        "---", "title: AI知识库资料导入说明", "type: ai-knowledge-library-index", "status: active", "---", "",
        "# AI知识库资料导入说明", "",
        "> 用户点名资料的优先吸收结果：[[03-重点吸收/00_重点资料索引|重点资料索引]]", "",
        f"- 顶层源文件：{summary['top_level_file_count']}",
        f"- 顶层唯一文件：{summary['top_level_unique_sha256_count']}",
        f"- 压缩包内文档：{summary['archive_document_member_count']}",
        f"- 可浏览唯一文档：{summary['canonical_document_count']}",
        f"- 重复来源：{summary['duplicate_source_count']}",
        f"- 宿析方法条目：{len(method_entries)}", "",
        "## 使用边界", "",
        "该库用于检索和人工分析；案例数值、合同、订单、名单、模板和旧资料均不自动成为当前酒店事实。",
        "缺少酒店、平台/数据源、业务日期、采集时间、质量状态和保存回读时，不进入收益决策或自动执行。",
        "压缩包中的脚本与 Skill 仅做静态文档提取，未安装、未执行。", "",
        "## 主题入口", "", *topic_index_links, "",
        "## 机器清单", "", "- [[_manifest.json]]", "- [[_manifest.csv]]", "",
    ]
    write_utf8(obsidian / "00_导入说明.md", "\n".join(summary_lines))
    write_utf8(obsidian / "01_资料总索引.md",
        "\n".join(["# 资料总索引", "", *[link for links in topic_links.values() for link in sorted(set(links))], ""]),
    )
    write_utf8(suxios_output / "README.md",
        "\n".join([
            "# AI知识库资料吸收包", "",
            f"来源：`{source}`", "",
            f"版本：`{SEED_VERSION}`", "",
            "- `source-manifest.json`：全部顶层源文件的哈希、类型、解析状态与边界。",
            "- `method-pack.json`：从来源目录中已成型的方法、专题与案例 Markdown 提取的可检索参考条目。",
            "- `priority-pack.json`：用户点名资料经证据校准后的重点条目；外来 Skill 仅静态审阅，未安装、未执行。",
            "- `integrated-model.json`：全量方法与重点资料的统一概念、指标、诊断、动作和黄金样例模型。",
            "- 所有条目均为 `reference_only`，不含当前酒店事实，不授权 OTA/PMS/企微写入。",
            "- 运行 `scripts/sync_ai_knowledge_library.php --persist` 才会写入当前宿析OS数据库；不带 `--persist` 仅校验。",
            "",
        ]),
    )

    result = {
        "status": "success" if not any(key.startswith(("extract_failed", "read_failed")) for key in status_counts) else "partial_success",
        "summary": summary,
        "method_entry_count": len(method_entries),
        "obsidian_output": str(obsidian),
        "suxios_output": str(suxios_output),
        "manifest_sha256": sha256_bytes(manifest_json.encode("utf-8")),
        "method_pack_sha256": sha256_bytes(method_pack_json.encode("utf-8")),
    }
    print(json.dumps({"result": result}, ensure_ascii=False), flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
